'''
Autor: Luiz Fernando Antonelli Galati
'''

'''
Este programa lê um arquivo .csv que contém o peso, a altura, a cor do cabelo e a cor do olho de 50 pessoas fictícias.
Em seguida, o programa:
(1) Cria um novo arquivo .xlsx com todos os dados do arquivo .csv (em outras palavras, ele converte o arquivo .csv em um
arquivo .xlsx, mantendo o arquivo original);
(2) Calcula as medidas de tendência central das variáveis quantitativas peso e altura e as escreve no novo arquivo;
(3) Constrói a tabela de contingência das variáveis qualitativas "Cor do cabelo" e "Cor do olho" e desenha essa tabela no
novo arquivo;
(4) Constrói o gráfico de dispersão das variáveis "peso" e "altura", calcula a correlação entre elas e imprime essas infor-
mações (o gráfico e a correlação) no novo arquivo.

O programa pode ser facilmente adaptado para ler muitas tabelas semelhantes e extrair delas as mesmas informações.
''' 


import pandas as pd
import openpyxl
import matplotlib.pyplot as plt


''' Converte o arquivo csv 'arquivo1' para o arquivo xlsx 'arquivo2'. '''

def converte_xlsx (arquivo1, arquivo2):
	csv_dataframe = pd.read_csv (arquivo1)
	excel_writer = pd.ExcelWriter (arquivo2)
	csv_dataframe.to_excel (excel_writer, index = False)
	excel_writer.close ()


''' Calcula as medidas de tendência central e o desvio-padrão dos valores assumidos pelas variáveis quantitativas 'Peso' e 'Altura' 
em 'dataframe'. Escreve os resultados em 'planilha'. '''

def tendenciaCentral (planilha, dataframe):
    planilha.cell(1, 7).value = "Peso"
    planilha.cell(1, 8).value = "Altura"
    planilha.cell(2, 6).value = "Média"
    planilha.cell(3, 6).value = "Mediana"
    planilha.cell(4, 6).value = "Primeiro quartil"
    planilha.cell(5, 6).value = "Terceiro quartil"
    planilha.cell(6, 6).value = "Desvio-padrão"

    mediaPeso = dataframe['Peso'].mean()
    quartisPeso = dataframe['Peso'].quantile([0.25, 0.5, 0.75])
    desvioPeso = dataframe['Peso'].std()

    mediaAltura = dataframe['Altura'].mean()
    quartisAltura= dataframe['Altura'].quantile([0.25, 0.5, 0.75])
    desvioAltura = dataframe['Altura'].std() 

    planilha.cell(2, 7).value = mediaPeso
    planilha.cell(3, 7).value = quartisPeso[0.5]
    planilha.cell(4, 7).value = quartisPeso[0.25]
    planilha.cell(5, 7).value = quartisPeso[0.75]
    planilha.cell(6, 7).value = desvioPeso

    planilha.cell(2, 8).value = mediaAltura
    planilha.cell(3, 8).value = quartisAltura[0.5]
    planilha.cell(4, 8).value = quartisAltura[0.25]
    planilha.cell(5, 8).value = quartisAltura[0.75]
    planilha.cell(6, 8).value = desvioAltura


''' Constrói a tabela de contingência dos valores assumidos pelas variáveis qualitativas 'Cor cabelo' e 'Cor olho' 
em 'planilha'. Desenha a tabela em 'planilha'. '''

def tabelaContingencia (planilha):
    planilha.cell(11, 6).value = "Cor cabelo"
    planilha.cell(10, 7).value = "castanho"
    planilha.cell(11, 7).value = "loiro"
    planilha.cell(12, 7).value = "preto"
    planilha.cell(13, 7).value = "Total"

    planilha.cell(8, 9).value = "Cor olho"
    planilha.cell(9, 8).value = "azul"
    planilha.cell(9, 9).value = "castanho"
    planilha.cell(9, 10).value = "verde"
    planilha.cell(9, 11).value = "Total"

    azul_castanho = 0
    castanho_castanho = 0
    verde_castanho = 0

    azul_loiro = 0
    castanho_loiro = 0
    verde_loiro = 0

    azul_preto = 0
    castanho_preto = 0
    verde_preto = 0

    i = 2
    while (i <= planilha.max_row):
        cor_olho = planilha.cell(i, 3).value
        cor_cabelo = planilha.cell(i, 4).value
        if (cor_olho == "azul" and cor_cabelo == "castanho"):
            azul_castanho = azul_castanho + 1
        elif (cor_olho == "castanho" and cor_cabelo == "castanho"):
            castanho_castanho = castanho_castanho + 1
        elif (cor_olho == "verde" and cor_cabelo == "castanho"):
            verde_castanho = verde_castanho + 1
        elif (cor_olho == "azul" and cor_cabelo == "loiro"):
            azul_loiro = azul_loiro + 1
        elif (cor_olho == "castanho" and cor_cabelo == "loiro"):
            castanho_loiro = castanho_loiro + 1
        elif (cor_olho == "verde" and cor_cabelo == "loiro"):
            verde_loiro = verde_loiro + 1
        elif (cor_olho == "azul" and cor_cabelo == "preto"):
            azul_preto = azul_preto + 1
        elif (cor_olho == "castanho" and cor_cabelo == "preto"):
            castanho_preto = castanho_preto + 1
        elif (cor_olho == "verde" and cor_cabelo == "preto"):
            verde_preto = verde_preto + 1
            
        i = i + 1

    planilha['H10'] = azul_castanho
    planilha['I10'] = castanho_castanho
    planilha['J10'] = verde_castanho

    planilha['H11'] = azul_loiro
    planilha['I11'] = castanho_loiro
    planilha['J11'] = verde_loiro

    planilha['H12'] = azul_preto
    planilha['I12'] = castanho_preto
    planilha['J12'] = verde_preto

    planilha['K10'] = azul_castanho + castanho_castanho + verde_castanho
    planilha['K11'] = azul_loiro + castanho_loiro + verde_loiro
    planilha['K12'] = azul_preto + castanho_preto + verde_preto

    planilha['H13'] = azul_castanho + azul_loiro + azul_preto
    planilha['I13'] = castanho_castanho + castanho_loiro + castanho_preto
    planilha['J13'] = verde_castanho + verde_loiro + verde_preto

    planilha['K13'] = planilha['H13'].value + planilha['I13'].value + planilha['J13'].value


''' Constrói o gráfico de dispersão (scatter-plot) dos valores assumidos pelas variáveis 'Peso' e 'Altura' presentes em 'dataframe'.
Desenha o gráfico em 'planilha'. '''

def plotaGraficoDispersao (planilha, dataframe):
    plt.scatter (dataframe['Peso'], dataframe['Altura'])
    plt.title ('Gráfico de Dispersão de Peso vs Altura')
    plt.savefig ('scatter_plot.png')   
    
    img = openpyxl.drawing.image.Image ('scatter_plot.png')
    planilha.add_image (img, 'F15')
    
    correlacao = dataframe['Peso'].corr(dataframe['Altura'])      
    planilha["F40"] = "Correlação (peso, altura)"
    planilha["G40"] = correlacao    


def main ():
    converte_xlsx ("dados_fenotipo.csv", "dados_fenotipo_em_excel.xlsx")

    arquivo = openpyxl.load_workbook ("dados_fenotipo_em_excel.xlsx")
    planilha = arquivo.worksheets[0]
    planilha.title = "Análise descritiva"
    df = pd.read_excel ('dados_fenotipo_em_excel.xlsx')
    tendenciaCentral (planilha, df)
    tabelaContingencia (planilha)    
    plotaGraficoDispersao (planilha, df)

    arquivo.save (filename = "dados_fenotipo_em_excel.xlsx")

main ()